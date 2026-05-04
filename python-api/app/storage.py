from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.styles import Border, Side

from .models import ProcessedPayment, RegisterPaymentResponse, ReportResponse

# Client-facing columns only — no internal/system fields
HEADERS = [
    "Fecha",
    "Hora",
    "Nombre",
    "Monto (S/.)",
    "Tipo",
    "N° Operación",
    "Estado",
]

# Column widths matching HEADERS
_COL_WIDTHS = [13, 8, 30, 13, 9, 18, 16]

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# Row alternating fill
_ROW_ALT_FILL = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")

# Estado colors
_ESTADO_FILLS = {
    "Registrado": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Duplicado": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Invalido": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Error OCR": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}

_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def payments_path() -> Path:
    return Path(os.getenv("PAYMENTS_XLSX_PATH", "/app/pagos/pagos.xlsx"))


def report_path() -> Path:
    base = payments_path().parent
    return base / "reporte.xlsx"


def normalize_query_date(value: str | None = None) -> str:
    if not value:
        return datetime.now().strftime("%d/%m/%Y")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return value


def _apply_header_style(sheet: Worksheet) -> None:
    for col_idx, (header, width) in enumerate(zip(HEADERS, _COL_WIDTHS), start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"


def _style_data_row(sheet: Worksheet, row_num: int, estado: str) -> None:
    """Apply borders, alternating fill, estado color, and number format to a data row."""
    use_alt = row_num % 2 == 0
    estado_fill = _ESTADO_FILLS.get(estado)
    num_cols = len(HEADERS)
    for col_idx in range(1, num_cols + 1):
        cell = sheet.cell(row=row_num, column=col_idx)
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        # Monto column (col 4): currency number format
        if col_idx == 4:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        # Estado column (col 7): colored background
        if col_idx == 7 and estado_fill:
            cell.fill = estado_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif use_alt and col_idx != 7:
            cell.fill = _ROW_ALT_FILL


def _worksheet() -> Worksheet:
    path = payments_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Pagos"
        _apply_header_style(sheet)
        workbook.save(path)
        return sheet

    workbook = load_workbook(path)
    sheet = workbook.active
    current_headers = [cell.value for cell in sheet[1]]
    if current_headers != HEADERS:
        # Schema changed — start fresh to avoid mismatched columns
        path.unlink()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Pagos"
        _apply_header_style(sheet)
        workbook.save(path)
    return sheet


def _save_sheet(sheet: Worksheet) -> None:
    sheet.parent.save(payments_path())


def payment_key(payment: ProcessedPayment) -> tuple[str, str, str]:
    return (
        (payment.tipo or "").strip().lower(),
        (payment.operacion or "").strip(),
        f"{payment.monto or 0:.2f}",
    )


# Column indices (0-based) in the 7-column schema:
# 0=Fecha, 1=Hora, 2=Nombre, 3=Monto, 4=Tipo, 5=N°Op, 6=Estado

def _row_key(row: tuple) -> tuple[str, str, str]:
    tipo = str(row[4] or "").strip().lower()
    operation = str(row[5] or "").strip()
    try:
        amount = f"{float(row[3] or 0):.2f}"
    except (TypeError, ValueError):
        amount = "0.00"
    return tipo, operation, amount


def is_duplicate(sheet: Worksheet, payment: ProcessedPayment) -> bool:
    key = payment_key(payment)
    if not all(key):
        return False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[6] or "") == "Registrado" and _row_key(row) == key:
            return True
    return False


def register_payment(payment: ProcessedPayment) -> RegisterPaymentResponse:
    sheet = _worksheet()
    if payment.valido and is_duplicate(sheet, payment):
        payment.estado = "Duplicado"
        return RegisterPaymentResponse(
            registrado=False,
            estado="Duplicado",
            mensaje="El pago ya habia sido registrado.",
            pago=payment,
            reporte=build_report(payment.fecha).model_dump(),
        )

    if not payment.valido and payment.estado == "Registrado":
        payment.estado = "Invalido"

    row_data = [
        payment.fecha or normalize_query_date(),
        payment.hora or datetime.now().strftime("%H:%M"),
        payment.nombre or "",
        payment.monto or 0,
        payment.tipo or "",
        payment.operacion or "",
        payment.estado,
    ]
    sheet.append(row_data)
    new_row_num = sheet.max_row
    _style_data_row(sheet, new_row_num, payment.estado)
    _save_sheet(sheet)
    report = build_report(payment.fecha)
    save_daily_report(report)
    return RegisterPaymentResponse(
        registrado=payment.estado == "Registrado",
        estado=payment.estado,
        mensaje="Pago registrado correctamente."
        if payment.estado == "Registrado"
        else "Pago guardado para revision.",
        pago=payment,
        reporte=report.model_dump(),
    )


def list_payments(fecha: str | None = None) -> list[dict]:
    sheet = _worksheet()
    target_date = normalize_query_date(fecha) if fecha else None
    rows: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(HEADERS, row, strict=False))
        if target_date and item.get("Fecha") != target_date:
            continue
        rows.append(item)
    return rows


def _amount_key(row: dict) -> float:
    """Return the numeric amount from a row regardless of header name."""
    for key in ("Monto (S/.)", "Monto"):
        val = row.get(key)
        if val is not None:
            try:
                return float(val)
            except (TypeError, ValueError):
                pass
    return 0.0


def build_report(fecha: str | None = None) -> ReportResponse:
    target_date = normalize_query_date(fecha)
    rows = list_payments(target_date)
    registered = [row for row in rows if row.get("Estado") == "Registrado"]

    total_yape = sum(_amount_key(row) for row in registered if row.get("Tipo") == "Yape")
    total_plin = sum(_amount_key(row) for row in registered if row.get("Tipo") == "Plin")
    return ReportResponse(
        fecha=target_date,
        total_recaudado=round(total_yape + total_plin, 2),
        cantidad_pagos=len(registered),
        total_yape=round(total_yape, 2),
        total_plin=round(total_plin, 2),
        duplicados=sum(1 for row in rows if row.get("Estado") == "Duplicado"),
        errores_ocr=sum(1 for row in rows if row.get("Estado") == "Error OCR"),
        invalidos=sum(1 for row in rows if row.get("Estado") == "Invalido"),
    )


# ---------------------------------------------------------------------------
# Daily report — separate Excel file
# ---------------------------------------------------------------------------

_REPORT_HEADERS = ["Fecha", "Total recaudado (S/.)", "Pagos Yape (S/.)", "Pagos Plin (S/.)", "Cantidad de pagos", "Duplicados"]
_REPORT_COL_WIDTHS = [14, 22, 18, 18, 20, 14]
_REPORT_HEADER_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")


def _report_worksheet() -> Worksheet:
    path = report_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Diario"
        _apply_report_header(ws)
        wb.save(path)
        return ws

    wb = load_workbook(path)
    ws = wb.active
    if [cell.value for cell in ws[1]] != _REPORT_HEADERS:
        path.unlink()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Diario"
        _apply_report_header(ws)
        wb.save(path)
    return ws


def _apply_report_header(ws: Worksheet) -> None:
    for col_idx, (header, width) in enumerate(zip(_REPORT_HEADERS, _REPORT_COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _REPORT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


_CURRENCY_FMT = '#,##0.00'
_NUM_COLS = {2, 3, 4}  # 1-based: Total, Yape, Plin

_TOTAL_FILL = PatternFill(start_color="1E3A1E", end_color="1E3A1E", fill_type="solid")
_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)


def _style_report_row(ws: Worksheet, row_num: int) -> None:
    use_alt = row_num % 2 == 0
    for col_idx in range(1, len(_REPORT_HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if col_idx in _NUM_COLS:
            cell.number_format = _CURRENCY_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if use_alt:
            cell.fill = _ROW_ALT_FILL


def _write_total_row(ws: Worksheet, data_start: int, data_end: int) -> None:
    """Write or overwrite the TOTAL row at data_end+1."""
    total_row = data_end + 1
    num_cols = len(_REPORT_HEADERS)

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        cell.border = _THIN_BORDER

        if col_idx == 1:
            cell.value = "TOTAL"
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        elif col_idx in _NUM_COLS:
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            cell.number_format = _CURRENCY_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[total_row].height = 22


def _remove_old_total_row(ws: Worksheet) -> int:
    """Remove any existing TOTAL row and return the last real data row number."""
    last_data = 1
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").upper() == "TOTAL":
            ws.delete_rows(row[0].row)
            break
        last_data = row[0].row
    return last_data


def save_daily_report(report: ReportResponse) -> None:
    """Upsert a row for today in reporte.xlsx, then refresh the TOTAL row at the bottom."""
    ws = _report_worksheet()
    target_date = report.fecha or normalize_query_date()

    # Remove existing TOTAL row before any changes
    _remove_old_total_row(ws)

    # Find existing row for this date and update it
    updated = False
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "") == target_date:
            row[0].value = target_date
            row[1].value = report.total_recaudado
            row[2].value = report.total_yape
            row[3].value = report.total_plin
            row[4].value = report.cantidad_pagos
            row[5].value = report.duplicados
            _style_report_row(ws, row[0].row)
            updated = True
            break

    if not updated:
        ws.append([
            target_date,
            report.total_recaudado,
            report.total_yape,
            report.total_plin,
            report.cantidad_pagos,
            report.duplicados,
        ])
        _style_report_row(ws, ws.max_row)

    # Write TOTAL row at the bottom
    last_data = ws.max_row
    _write_total_row(ws, data_start=2, data_end=last_data)
    ws.parent.save(report_path())
